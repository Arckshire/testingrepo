import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO, StringIO
from openpyxl import Workbook
import csv

# === Column name configuration (change these if your raw file uses different labels) ===
TRACKED_COL_NAME = "Tracked"
PICKUP_TS_COL_NAME = "Pickup Departure UTC Timestamp Raw"
DROPOFF_TS_COL_NAME = "Drop-off Arrival UTC Timestamp Raw"

BOL_COL_NAME = "Bill of lading"
PICKUP_NAME_COL_NAME = "Pickup name"
PICKUP_CITY_STATE_COL_NAME = "Pickup cityState"
PICKUP_COUNTRY_COL_NAME = "Pickup country"
DROPOFF_NAME_COL_NAME = "Dropoff name"
DROPOFF_CITY_STATE_COL_NAME = "Dropoff city state"
DROPOFF_COUNTRY_COL_NAME = "Dropoff country"


def normalize_col_name(name: str) -> str:
    """Normalize column names to compare them case-insensitively and ignoring spaces/punctuation."""
    return "".join(ch.lower() for ch in str(name) if ch.isalnum())


def find_column(df: pd.DataFrame, target_name: str) -> str:
    """Return the actual column name from df that matches target_name (case/space insensitive)."""
    norm_target = normalize_col_name(target_name)
    for col in df.columns:
        if normalize_col_name(col) == norm_target:
            return col
    raise KeyError(f"Could not find a column matching '{target_name}' in the uploaded file.")


def coerce_tracked_flag(series: pd.Series) -> pd.Series:
    """Convert the 'tracked' column into a clean boolean Series."""
    if series.dtype == bool:
        return series.fillna(False)

    # Handle booleans stored as strings / numbers
    s = series.astype(str).str.strip().str.lower()
    truthy = {"true", "1", "yes", "y"}
    return s.isin(truthy)


def clean_and_parse_timestamp(series: pd.Series) -> pd.Series:
    """Clean obvious 'zero' placeholders and parse timestamps to datetime (UTC)."""
    s = series.replace({0: pd.NA, 0.0: pd.NA, "0": pd.NA, "0000-00-00 00:00:00": pd.NA})
    return pd.to_datetime(s, errors="coerce", utc=True)


def round_transit_days(days: float) -> float:
    """Apply custom rounding logic for transit days."""
    if pd.isna(days):
        return np.nan
    if days <= 0:
        # Non-positive durations are considered missed milestones
        return np.nan
    if days < 0.5:
        return 0.5
    if days < 1:
        return 1.0

    integer_part = int(np.floor(days))
    frac_part = float(days - integer_part)
    if frac_part < 0.5:
        return float(integer_part)
    else:
        return float(integer_part + 1)


def build_detail_df(
    df_tracked: pd.DataFrame,
    pickup_ts: pd.Series,
    dropoff_ts: pd.Series,
    rounded_transit: pd.Series,
    missed_mask: pd.Series,
) -> pd.DataFrame:
    """Build the detailed table for tracked shipments with valid in-transit time."""
    valid_mask = ~missed_mask & rounded_transit.notna()

    df_valid = df_tracked.loc[valid_mask].copy()
    df_valid["In transit time"] = rounded_transit.loc[valid_mask].values

    # Map raw column names to clean output labels
    bol_col = find_column(df_tracked, BOL_COL_NAME)
    pickup_name_col = find_column(df_tracked, PICKUP_NAME_COL_NAME)
    pickup_city_state_col = find_column(df_tracked, PICKUP_CITY_STATE_COL_NAME)
    pickup_country_col = find_column(df_tracked, PICKUP_COUNTRY_COL_NAME)
    dropoff_name_col = find_column(df_tracked, DROPOFF_NAME_COL_NAME)
    dropoff_city_state_col = find_column(df_tracked, DROPOFF_CITY_STATE_COL_NAME)
    dropoff_country_col = find_column(df_tracked, DROPOFF_COUNTRY_COL_NAME)

    detail_df = pd.DataFrame({
        "Bill of lading": df_valid[bol_col].values,
        "Pickup name": df_valid[pickup_name_col].values,
        "Pickup City State": df_valid[pickup_city_state_col].values,
        "Pickup country": df_valid[pickup_country_col].values,
        "Dropoff name": df_valid[dropoff_name_col].values,
        "Dropoff City State": df_valid[dropoff_city_state_col].values,
        "Dropoff country": df_valid[dropoff_country_col].values,
        "In transit time": df_valid["In transit time"].values,
    })

    return detail_df


def build_excel_file(summary_counts: dict, detail_df: pd.DataFrame) -> bytes:
    """Create the processed Excel file with the required layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "In-Transit Report"

    # Summary header row
    ws["A1"] = "Label"
    ws["B1"] = "Shipment Count"
    ws["D1"] = "Definition of in transit time"
    ws["E1"] = "Time taken from departure to arrival"

    # Summary rows
    ws["A2"] = "Tracked"
    ws["B2"] = summary_counts["tracked_count"]

    ws["A3"] = "Missed milestone"
    ws["B3"] = summary_counts["missed_milestone_count"]

    ws["A4"] = "Untracked"
    ws["B4"] = summary_counts["untracked_count"]

    ws["A5"] = "Grand total"
    ws["B5"] = summary_counts["grand_total"]

    # Row 6 left blank (separator)

    # Detailed table header at row 7
    # Using contiguous columns A–H for the detail table.
    headers = [
        "Bill of lading",
        "Pickup name",
        "Pickup City State",
        "Pickup country",
        "Dropoff name",
        "Dropoff City State",
        "Dropoff country",
        "In transit time",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=7, column=col_idx, value=header)

    # Data from row 8 downward
    start_row = 8
    for row_idx, (_, row) in enumerate(detail_df.iterrows(), start=start_row):
        ws.cell(row=row_idx, column=1, value=row["Bill of lading"])
        ws.cell(row=row_idx, column=2, value=row["Pickup name"])
        ws.cell(row=row_idx, column=3, value=row["Pickup City State"])
        ws.cell(row=row_idx, column=4, value=row["Pickup country"])
        ws.cell(row=row_idx, column=5, value=row["Dropoff name"])
        ws.cell(row=row_idx, column=6, value=row["Dropoff City State"])
        ws.cell(row=row_idx, column=7, value=row["Dropoff country"])
        ws.cell(row=row_idx, column=8, value=row["In transit time"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_csv_file(summary_counts: dict, detail_df: pd.DataFrame) -> str:
    """Create a CSV string approximating the Excel layout."""
    output = StringIO()
    writer = csv.writer(output)

    # Summary section
    writer.writerow(["Label", "Shipment Count", "", "Definition of in transit time", "Time taken from departure to arrival"])
    writer.writerow(["Tracked", summary_counts["tracked_count"]])
    writer.writerow(["Missed milestone", summary_counts["missed_milestone_count"]])
    writer.writerow(["Untracked", summary_counts["untracked_count"]])
    writer.writerow(["Grand total", summary_counts["grand_total"]])

    # Blank row
    writer.writerow([])

    # Detailed header
    writer.writerow([
        "Bill of lading",
        "Pickup name",
        "Pickup City State",
        "Pickup country",
        "Dropoff name",
        "Dropoff City State",
        "Dropoff country",
        "In transit time",
    ])

    # Detailed rows
    for _, row in detail_df.iterrows():
        writer.writerow([
            row["Bill of lading"],
            row["Pickup name"],
            row["Pickup City State"],
            row["Pickup country"],
            row["Dropoff name"],
            row["Dropoff City State"],
            row["Dropoff country"],
            row["In transit time"],
        ])

    return output.getvalue()


def main():
    st.title("FTL In-Transit Time Calculator")

    st.write(
        "Upload a raw FTL tracking file (CSV or Excel). "
        "The app will classify tracked/untracked shipments, "
        "detect missed milestones, and calculate in-transit time in days."
    )

    uploaded_file = st.file_uploader(
        "Upload raw FTL tracking file",
        type=["csv", "xlsx", "xls"],
        help="CSV, XLSX, or XLS files are supported."
    )

    if uploaded_file is None:
        st.info("👆 Start by uploading a file.")
        return

    # Read the file into a DataFrame
    try:
        if uploaded_file.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Could not read the uploaded file: {e}")
        return

    if df.empty:
        st.warning("The uploaded file appears to be empty.")
        return

    # Map column names (case/space-insensitive)
    try:
        tracked_col = find_column(df, TRACKED_COL_NAME)
        pickup_ts_col = find_column(df, PICKUP_TS_COL_NAME)
        dropoff_ts_col = find_column(df, DROPOFF_TS_COL_NAME)
    except KeyError as e:
        st.error(str(e))
        return

    # Step 1 – Tracked vs Untracked
    tracked_flag = coerce_tracked_flag(df[tracked_col])
    df_tracked = df[tracked_flag].copy()
    df_untracked = df[~tracked_flag].copy()

    tracked_count = len(df_tracked)
    untracked_count = len(df_untracked)
    total_rows = len(df)

    # Step 2 – In-transit time and missed milestones (within tracked)
    pickup_ts = clean_and_parse_timestamp(df_tracked[pickup_ts_col])
    dropoff_ts = clean_and_parse_timestamp(df_tracked[dropoff_ts_col])

    transit_timedelta = dropoff_ts - pickup_ts
    transit_days = transit_timedelta.dt.total_seconds() / (24 * 60 * 60)

    # Missed milestones: missing timestamps or non-positive duration
    missed_mask = pickup_ts.isna() | dropoff_ts.isna() | (transit_days <= 0)

    # Round valid transit days using your custom logic
    rounded_transit = transit_days.apply(round_transit_days)

    missed_milestone_count = int(missed_mask.sum())
    actual_tracked_with_valid_transit = int((~missed_mask & rounded_transit.notna()).sum())

    # We treat grand total as the total number of rows in the original file
    grand_total = total_rows

    summary_counts = {
        "tracked_count": tracked_count,
        "missed_milestone_count": missed_milestone_count,
        "untracked_count": untracked_count,
        "grand_total": grand_total,
        "actual_tracked_with_valid_transit": actual_tracked_with_valid_transit,
    }

    # Build detail table
    try:
        detail_df = build_detail_df(
            df_tracked=df_tracked,
            pickup_ts=pickup_ts,
            dropoff_ts=dropoff_ts,
            rounded_transit=rounded_transit,
            missed_mask=missed_mask,
        )
    except KeyError as e:
        st.error(str(e))
        return

    # Display summary
    st.subheader("Summary")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Tracked", tracked_count)
    c2.metric("Missed milestone", missed_milestone_count)
    c3.metric("Untracked", untracked_count)
    c4.metric("Grand total", grand_total)

    st.caption(
        "Note: 'Tracked' includes all shipments where the tracked flag is TRUE "
        "(including those classified as missed milestones). "
        "'Grand total' is the total number of rows in the original file."
    )

    # Display detailed table
    st.subheader("Tracked shipments with valid in-transit time (days)")
    st.dataframe(detail_df, use_container_width=True)

    # Build output files
    excel_bytes = build_excel_file(summary_counts, detail_df)
    csv_text = build_csv_file(summary_counts, detail_df)

    st.subheader("Download processed data")

    st.download_button(
        label="⬇️ Download Excel report",
        data=excel_bytes,
        file_name="in_transit_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="⬇️ Download CSV report",
        data=csv_text.encode("utf-8"),
        file_name="in_transit_report.csv",
        mime="text/csv",
    )


if __name__ == "__main__":
    main()
